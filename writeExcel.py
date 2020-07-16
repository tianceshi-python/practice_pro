# encoding:utf-8

import os
import getpathInfo   #导入自已定义的内部类，该类返回项目的绝对路径
#调用读取Excel的第三方库
from xlwt import *
from openpyxl import Workbook,load_workbook
from common import logPrintClass




#拿到项目的绝对路径
path = getpathInfo.get_Path()

class writeExcel:

    def __init__(self,caseDirName1,caseDirName2,xls_name):

        self.log = logPrintClass.Log()

        self.caseDirName1 = caseDirName1
        self.caseDirName2 = caseDirName2
        self.xls_name = xls_name

        self.path = os.path.join(path,self.caseDirName1,self.caseDirName2)

    def write_xml(self,sheet_name,case_name, message):

        self.log.debug('write_xml: data write into xls start.....')

        os.chdir(self.path)
        #filename = os.path.join(path, caseDirName1, caseDirName2, xls_name)
        if not os.path.exists(self.xls_name):
            mywb = Workbook(encoding='UTF-8')
            mysheet = mywb.create_sheet(sheet_name)
        else:
            mywb = load_workbook(self.xls_name)        #注意：直接操作文件名
            mysheet = mywb[sheet_name]

        rows = mysheet.max_row  # 获取行数
        #cols = mysheet.max_column  # 获取列数
        #print('rows is: ',rows)
        #print('cols is: ',cols)

        flag = True         #作为表中是否有case_name的标志位
        # 如果表中有这个case的结果记录，则更新返回数据的body，且将标志位置为false
        for i in range(rows+1):
            i = i + 1
            if mysheet.cell(i,1).value == case_name:
                mysheet.cell(i, 2).value = message
                flag = False
                self.log.debug('write_xml: update body data....')

        # 在表中没有该条case的返回body记录，则在表中增加该case的名称和返回消息体
        if flag:
            mysheet.cell(rows + 1, 1).value = case_name
            mysheet.cell(rows + 1, 2).value = message
            self.log.debug('write_xml: no case_name,The new writing case_name and body...')
        else:
            pass

        #写书数据后保存该文件
        mywb.save(self.xls_name)    #注意：直接操作文件名
        mywb.close()





if  __name__ == "__main__":
    obj = writeExcel('testCase','casedata','conferenceControl_casedate.xlsx')
    obj.write_xml(sheet_name = 'responseData',case_name = 'test_QueryMeetingStatus001', message = 'rrrrr')